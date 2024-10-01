from pymongo import MongoClient
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


client = MongoClient('mongodb+srv://kanderpyfish:Boston2011%23@logininfo.s1rjc4g.mongodb.net/?retryWrites=true&w=majority&appName=LoginInfo')  # Replace with your MongoDB connection string
db = client['LoginInfo']  # Replace 'your_database' with your actual database name
collection = db['UserInfo']  # Replace 'your_collection' with your actual collection name
company_name = "Dell"

documents1 = collection.find({
    "form_type": "adminFormData",
    "company": company_name
})

cursor = collection.find({
    "form_type": "salesRepFormData",
    "company": company_name
}) 

adminFormData = documents1[0] if documents1 else None
quarterly_splits = None
tcv = None
acv = None

if adminFormData:
    result_data = adminFormData.get('result', {})
    quarterly_splits = result_data.get('Quarterly Splits')
    tcv = result_data.get('TCV')
    acv = result_data.get('ACV')
    a1 = adminFormData.get('a1')
    a2 = adminFormData.get('a2')
    a3 = adminFormData.get('a3')
    a4 = adminFormData.get('a4')

# Retrieve all documents in the collection

def is_merged_cell(ws, cell):
    """Check if a cell is part of a merged cell range."""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False

def auto_adjust_columns(ws):
    """Adjust the column width if it's not part of a merged cell."""
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if ws.cell(cell.row, cell.column).value is not None:
                if not is_merged_cell(ws, cell):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
        column_letter = get_column_letter(col[0].column)
        ws.column_dimensions[column_letter].width = max_length + 2





        
#Create sample excel sheet
def create_excel(documents, quarterly_splits, tcv, acv,a1,a2,a3,a4 ):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define styles
    title_font = Font(size=14, bold=True)
    section_font = Font(size=11, bold=True)
    normal_font = Font(size=11)
    center_aligned_text = Alignment(horizontal="left")

    # Add title
    ws.append(["2024 Fiscal Year Sales Incentive Compensation Plan"])
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')

    # Add effective date
    ws.append(["Effective Jan 1, 2024"])
    ws['A2'].font = section_font
    ws.merge_cells('A2:J2')

    # Add introduction
    ws.append([""])
    ws.append(["Endless Moments Inc. is delighted to present the comprehensive compensation plan outlined below for your consideration:"])
    ws['A4'].font = normal_font
    ws.merge_cells('A4:J4')

    # Add Employee details
    ws.append(["Employee Name", "Title", "Base Salary", "Commission Rate"])
    for document in documents:
        employee_name = document.get('employeeName', 'N/A')
        title = document.get('title', 'N/A')
        base_salary = document.get('baseSalary', 'N/A')
        com_rate = f"{document.get('comRate', 'N/A')}%"  # Assuming comRate is a percentage

        ws.append([employee_name, title, f"${base_salary}", com_rate])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_aligned_text
            cell.font = normal_font
    
    # Spacer row
    ws.append([""])

    # Add Individual Goals table
    row = ["2024 quota"]
    product_type1 = document.get("productType1", "N/A")
    product_type2 = document.get("productType2", "N/A")
    product_type3 = document.get("productType3", "N/A")
    hardware_value = safe_float_conversion(document.get("hardwareQuota", "0"))
    software_value = safe_float_conversion(document.get("softwareQuota", "0"))
    services_value = safe_float_conversion(document.get("servicesQuota", "0"))
    total_value = safe_float_conversion(document.get("totalQuota", "0"))
    if (product_type1=="on" and hardware_value != 0):
        row.extend(["Hardware"])
    if (product_type2=="on" and software_value != 0):
        row.extend(["Software"])
    if (product_type3=="on" and services_value != 0):
        row.extend(["Services"])
    row.extend(["Total"])
    ws.append(row)

    row2=["Annual Quota"]
    if (product_type1=="on" and hardware_value != 0):
        row2.extend([hardware_value])
    if (product_type2=="on" and software_value != 0):
        row2.extend([software_value])
    if (product_type3=="on" and services_value != 0):
        row2.extend([services_value])
    row2.extend([total_value])
    ws.append(row2)

    row3=["Q1"]
    if (product_type1=="on" and hardware_value != 0):   
        row3.extend([hardware_value * (quarterly_splits[0]/100)])
    if (product_type2=="on" and software_value != 0):
        row3.extend([software_value * (quarterly_splits[0]/100)])
    if (product_type3=="on" and services_value != 0):
        row3.extend([services_value * (quarterly_splits[0]/100)])
    row3.extend([total_value * (quarterly_splits[0]/100)])
    ws.append(row3)

    row4=["Q2"]
    if (product_type1=="on" and hardware_value != 0):   
        row4.extend([hardware_value * (quarterly_splits[1]/100)])
    if (product_type2=="on" and software_value != 0):
        row4.extend([software_value * (quarterly_splits[1]/100)])
    if (product_type3=="on" and services_value != 0):
        row4.extend([services_value * (quarterly_splits[1]/100)])
    row4.extend([total_value * (quarterly_splits[1]/100)])
    ws.append(row4) 

    row5=["Q3"]
    if (product_type1=="on" and hardware_value != 0):   
        row5.extend([hardware_value * (quarterly_splits[2]/100)])
    if (product_type2=="on" and software_value != 0):
        row5.extend([software_value * (quarterly_splits[2]/100)])
    if (product_type3=="on" and services_value != 0):
        row5.extend([services_value * (quarterly_splits[2]/100)])
    row5.extend([total_value * (quarterly_splits[2]/100)])
    ws.append(row5)
    
    row6=["Q4"]
    if (product_type1=="on" and hardware_value != 0):   
        row6.extend([hardware_value * (quarterly_splits[3]/100)])
    if (product_type2=="on" and software_value != 0):
        row6.extend([software_value * (quarterly_splits[3]/100)])
    if (product_type3=="on" and services_value != 0):
        row6.extend([services_value * (quarterly_splits[3]/100)])
    row6.extend([total_value * (quarterly_splits[3]/100)])
    ws.append(row6) 
    
    # Spacer row
    ws.append([""])
    ws.append(["Commission rates", "0-50%", "50-99%", "100-149%", "150%+"])
    row7=["Accelerators",a1,a2,a3,a4]
    ws.append(row7)


    # Add Commission Rates table


    # Spacer rows
    ws.append([""])
    ws.append([""])

    # Add Confirmation statement
    ws.append(["I confirm that I have reviewed and comprehended the 2021 Sales Compensation Plan, including my designated quota and region assignments, as well as the commission calculation methodology."])
    ws['A26'].font = normal_font
    ws.merge_cells('A26:J26')

    # Add placeholders for Employee Name and Signature
    ws.append(["Employee Name"])
    ws.append(["Employee Signature"])

    # Apply alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_aligned_text
            
    auto_adjust_columns(ws)

    # Save the workbook to a file
    excel_path = r'C:\Users\suvid\Downloads\New_Incentive_Plan.xlsx'
    wb.save(excel_path)
    title_font = Font(size=12, bold=True)
    header_font = Font(size=11, bold=True)
    content_font = Font(size=11)

    for cell in ws["A"]: # Assuming title is in column A
        cell.font = title_font
        cell.alignment = Alignment(horizontal="left")

    for cell in ws[1]: # Assuming header is in row 1
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Apply content font and alignment for all other cells
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.font = content_font
            cell.alignment = Alignment(horizontal="left")

    return excel_path

# You would call this function when you need to create the Excel file.
def safe_float_conversion(value, default=0.0):
    try:
        return float(value)
    except ValueError:
        return default

documents = list(cursor)
excel_file_path = create_excel(documents, quarterly_splits, tcv, acv,a1,a2,a3,a4)
print(f"Excel file created at: {excel_file_path}")
client.close()